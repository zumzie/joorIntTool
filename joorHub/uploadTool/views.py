from django.http import HttpResponse
from .forms import UploadFileForm
from django.shortcuts import render
import os
import openpyxl
import csv
import json
from .forms import UploadFileForm
from django.shortcuts import render, HttpResponse

def upload_home(request):
    return render(request, 'uploadTool/csvTool.html')


def upload_file_view(request):
    preset_headers = ["style_name", "style_color", "style_size", "style_description", "style_category", "style_country_of_origin"]  # Defined here for use in both GET and POST
    
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            filename, file_extension = os.path.splitext(uploaded_file.name)
            data = extract_data_from_file(uploaded_file, file_extension)
            return render(request, 'uploadTool/display_data.html', {'data': data, 'preset_headers': preset_headers})
        else:
            return HttpResponse("Form is not valid. Please upload a file.")
    else:
        form = UploadFileForm()
        return render(request, 'uploadTool/csvTool.html', {'form': form, 'preset_headers': preset_headers})



def extract_data_from_file(uploaded_file, file_extension):
    data = []
    if file_extension.lower() == '.csv':
        lines = uploaded_file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(lines)
        data = [row for row in reader]
    elif file_extension.lower() in ['.xlsx', '.xls']:
        workbook = openpyxl.load_workbook(uploaded_file)
        worksheet = workbook.active
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        data = [
            {headers[i]: cell.value for i, cell in enumerate(row)}
            for row in worksheet.iter_rows(min_row=2)
        ]
    print("Extracted Data:", data)
    return data

def convert_data_to_json(data):
    return json.dumps(data, indent=4)